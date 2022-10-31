import requests
from bs4 import BeautifulSoup
import pandas as pd 
import matplotlib.pyplot as plt
from pydoc import plain
import openpyxl

HEADERS = {
'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64; Trident/7.0; Touch; MAARJS; rv:11.0) like Gecko',
'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
'Accept-Encoding':'gzip,deflate,sdch',
'Accept-Language':'pt-BR; q=0.9;en-US,en;q=0.8',
'Connection':'Keep-alive',
'Referer':'https://www.google.com.br',
}


resposta = requests.get('https://www.zoom.com.br/search?q=sete%20dias%20sem%20fim&refinements%5B0%5D%5Bid%5D=categoryId&refinements%5B0%5D%5Bvalues%5D%5B0%5D=444', headers=HEADERS)

#print(resposta.text)#


soup = BeautifulSoup(resposta.text, 'html.parser')
lista = []
h4 = soup.find_all(class_="Text_Text__bOTfK Text_MobileHeadingS__XS_Au")
for i in h4:
    lista.append(i.text)

print(lista)

with open("lista.txt", "w") as arquivo:
    arquivo.write(str(lista))

df = pd.read_table('lista.txt')
df.to_excel('lista_excel.xlsx',"Valores")



planilha = openpyxl.Workbook()
planilha.create_sheet('Sete dias sem fim')


aba_produtos = planilha['Sete dias sem fim']
aba_produtos.append(['Preço', 'Loja'])
aba_produtos.append(['34,90', 'Shoptime'])
aba_produtos.append(['34,90', 'Americanas'])
aba_produtos.append(['32,45', 'Submarino'])

planilha.save('sdsf.xlsx')
planilha = openpyxl.load_workbook('sdsf.xlsx')

plt.bar('Preço', 'Loja', facecolor='red')
plt.title('Grupos x Quantidades')
plt.savefig('grafico1.png', dpi=200)
plt.show